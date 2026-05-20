from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import json
import os
import base64
import random
import sqlite3
from datetime import datetime
import hashlib
import re
import urllib.error
import urllib.parse
import urllib.request
from io import BytesIO

try:
    from openpyxl import load_workbook
    from openpyxl import Workbook
except Exception:
    load_workbook = None
    Workbook = None

app = Flask(__name__)
CORS(app)

DB_PATH = 'words.db'


def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def table_columns(cursor, table_name):
    cursor.execute(f"PRAGMA table_info({table_name})")
    return {row[1] for row in cursor.fetchall()}


def has_unique_constraint(cursor, table_name, expected_columns):
    expected = list(expected_columns)
    cursor.execute(f"PRAGMA index_list({table_name})")
    indexes = cursor.fetchall()
    for idx in indexes:
        # PRAGMA index_list columns: seq, name, unique, origin, partial
        if len(idx) < 3 or int(idx[2]) != 1:
            continue
        idx_name = idx[1]
        cursor.execute(f"PRAGMA index_info({idx_name})")
        cols = [row[2] for row in cursor.fetchall()]
        if cols == expected:
            return True
    return False


def ensure_schema():
    conn = get_conn()
    cursor = conn.cursor()

    cursor.execute(
        '''
        CREATE TABLE IF NOT EXISTS books (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL DEFAULT 0,
            name TEXT NOT NULL,
            tag TEXT DEFAULT '',
            description TEXT DEFAULT '',
            cover TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        '''
    )

    cursor.execute(
        '''
        CREATE TABLE IF NOT EXISTS words (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL DEFAULT 0,
            word TEXT,
            meaning TEXT,
            pos TEXT DEFAULT '未知',
            phonetic TEXT DEFAULT '',
            audio TEXT DEFAULT '',
            book_id INTEGER
        )
        '''
    )

    cursor.execute(
        '''
        CREATE TABLE IF NOT EXISTS wrong_words (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL DEFAULT 0,
            word TEXT NOT NULL,
            meaning TEXT NOT NULL,
            pos TEXT DEFAULT '未知',
            error_count INTEGER NOT NULL DEFAULT 1,
            is_mastered INTEGER NOT NULL DEFAULT 0,
            is_focus INTEGER NOT NULL DEFAULT 0,
            needs_review INTEGER NOT NULL DEFAULT 1,
            last_wrong_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, word, meaning)
        )
        '''
    )

    cursor.execute(
        '''
        CREATE TABLE IF NOT EXISTS study_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL DEFAULT 0,
            word TEXT NOT NULL,
            book_id INTEGER,
            action TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        '''
    )
    cursor.execute(
        '''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            email TEXT NOT NULL UNIQUE,
            avatar TEXT DEFAULT '',
            password_hash TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        '''
    )
    cursor.execute(
        '''
        CREATE TABLE IF NOT EXISTS readings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL DEFAULT 0,
            title TEXT NOT NULL,
            category TEXT NOT NULL DEFAULT '英文短文',
            summary TEXT DEFAULT '',
            english TEXT NOT NULL,
            chinese TEXT NOT NULL,
            date TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        '''
    )
    reading_columns = table_columns(cursor, 'readings')
    if 'user_id' not in reading_columns:
        cursor.execute("ALTER TABLE readings ADD COLUMN user_id INTEGER NOT NULL DEFAULT 0")

    word_columns = table_columns(cursor, 'words')
    if 'pos' not in word_columns:
        cursor.execute("ALTER TABLE words ADD COLUMN pos TEXT DEFAULT '未知'")
    if 'phonetic' not in word_columns:
        cursor.execute("ALTER TABLE words ADD COLUMN phonetic TEXT DEFAULT ''")
    if 'audio' not in word_columns:
        cursor.execute("ALTER TABLE words ADD COLUMN audio TEXT DEFAULT ''")
    if 'book_id' not in word_columns:
        cursor.execute("ALTER TABLE words ADD COLUMN book_id INTEGER")
    if 'user_id' not in word_columns:
        cursor.execute("ALTER TABLE words ADD COLUMN user_id INTEGER NOT NULL DEFAULT 0")
    book_columns = table_columns(cursor, 'books')
    if 'tag' not in book_columns:
        cursor.execute("ALTER TABLE books ADD COLUMN tag TEXT DEFAULT ''")
    if 'user_id' not in book_columns:
        cursor.execute("ALTER TABLE books ADD COLUMN user_id INTEGER NOT NULL DEFAULT 0")

    log_columns = table_columns(cursor, 'study_logs')
    if 'book_id' not in log_columns:
        cursor.execute("ALTER TABLE study_logs ADD COLUMN book_id INTEGER")
    if 'user_id' not in log_columns:
        cursor.execute("ALTER TABLE study_logs ADD COLUMN user_id INTEGER NOT NULL DEFAULT 0")
    user_columns = table_columns(cursor, 'users')
    if 'avatar' not in user_columns:
        cursor.execute("ALTER TABLE users ADD COLUMN avatar TEXT DEFAULT ''")

    wrong_columns = table_columns(cursor, 'wrong_words')
    if 'user_id' not in wrong_columns:
        cursor.execute("ALTER TABLE wrong_words ADD COLUMN user_id INTEGER NOT NULL DEFAULT 0")
        wrong_columns = table_columns(cursor, 'wrong_words')
    legacy_wrong = not {'error_count', 'is_mastered', 'is_focus', 'needs_review', 'last_wrong_at', 'pos'}.issubset(
        wrong_columns
    )
    wrong_unique_ok = has_unique_constraint(cursor, 'wrong_words', ['user_id', 'word', 'meaning'])
    if legacy_wrong or not wrong_unique_ok:
        cursor.execute("ALTER TABLE wrong_words RENAME TO wrong_words_legacy")
        cursor.execute(
            '''
            CREATE TABLE wrong_words (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL DEFAULT 0,
                word TEXT NOT NULL,
                meaning TEXT NOT NULL,
                pos TEXT DEFAULT '未知',
                error_count INTEGER NOT NULL DEFAULT 1,
                is_mastered INTEGER NOT NULL DEFAULT 0,
                is_focus INTEGER NOT NULL DEFAULT 0,
                needs_review INTEGER NOT NULL DEFAULT 1,
                last_wrong_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(user_id, word, meaning)
            )
            '''
        )
        legacy_columns = table_columns(cursor, 'wrong_words_legacy')
        if {'user_id', 'error_count', 'is_mastered', 'is_focus', 'needs_review', 'last_wrong_at', 'pos'}.issubset(
            legacy_columns
        ):
            cursor.execute(
                '''
                INSERT INTO wrong_words (user_id, word, meaning, pos, error_count, is_mastered, is_focus, needs_review, last_wrong_at)
                SELECT user_id, word, meaning,
                       COALESCE(NULLIF(TRIM(pos), ''), '未知'),
                       COALESCE(error_count, 1),
                       COALESCE(is_mastered, 0),
                       COALESCE(is_focus, 0),
                       COALESCE(needs_review, 1),
                       COALESCE(last_wrong_at, CURRENT_TIMESTAMP)
                FROM wrong_words_legacy
                '''
            )
        else:
            cursor.execute(
                '''
                INSERT INTO wrong_words (user_id, word, meaning, pos, error_count, is_mastered, is_focus, needs_review, last_wrong_at)
                SELECT 0, word, meaning, '未知', COUNT(*), 0, 0, 1, MAX(CURRENT_TIMESTAMP)
                FROM wrong_words_legacy
                GROUP BY word, meaning
                '''
            )
        cursor.execute("DROP TABLE wrong_words_legacy")

    cursor.execute("SELECT COUNT(*) AS c FROM books")
    if (cursor.fetchone()['c'] or 0) == 0:
        cursor.execute(
            '''
            INSERT INTO books (name, description, cover)
            VALUES (?, ?, ?)
            ''',
            (
                '默认词书',
                '系统默认词书',
                'https://images.unsplash.com/photo-1456513080510-7bf3a84b82f8?auto=format&fit=crop&w=600&q=80',
            ),
        )

    cursor.execute("SELECT id FROM books ORDER BY id ASC LIMIT 1")
    default_book_id = cursor.fetchone()['id']
    cursor.execute("UPDATE words SET book_id = ? WHERE book_id IS NULL", (default_book_id,))

    cursor.execute("SELECT COUNT(*) AS c FROM readings")
    if (cursor.fetchone()['c'] or 0) == 0:
        today = datetime.now().strftime('%Y-%m-%d')
        seed_rows = [
            (
                'The City Mouse and the Country Mouse',
                '英文短文',
                'A classic short story about different lifestyles and choices.',
                'Once there were two mice. One lived in the city and the other lived in the country.',
                '从前有两只老鼠，一只住在城市，一只住在乡下。',
                today,
            ),
            (
                'Forrest Gump: Life Is Like a Box of Chocolates',
                '电影名句',
                'A famous quote about uncertainty and courage.',
                'Life is like a box of chocolates. You never know what you are going to get.',
                '生活就像一盒巧克力，你永远不知道下一颗是什么味道。',
                today,
            ),
        ]
        cursor.executemany(
            '''
            INSERT INTO readings (title, category, summary, english, chinese, date)
            VALUES (?, ?, ?, ?, ?, ?)
            ''',
            seed_rows,
        )

    conn.commit()
    conn.close()


def get_word_pos(cursor, word, meaning, user_id=None):
    cursor.execute(
        '''
        SELECT COALESCE(NULLIF(TRIM(pos), ''), '未知') AS pos
        FROM words
        WHERE word = ? AND meaning = ? {user_filter}
        LIMIT 1
        '''.format(user_filter='AND user_id = ?' if user_id else ''),
        (word, meaning, user_id) if user_id else (word, meaning),
    )
    row = cursor.fetchone()
    return row['pos'] if row else '未知'


def serialize_wrong(row):
    return {
        'word': row['word'],
        'meaning': row['meaning'],
        'pos': row['pos'] or '未知',
        'audio': row['audio'] or '',
        'bookName': row['book_name'] or '未知词书',
        'errorCount': row['error_count'],
        'isMastered': bool(row['is_mastered']),
        'isFocus': bool(row['is_focus']),
        'needsReview': bool(row['needs_review']),
        'lastWrongAt': row['last_wrong_at'],
    }


ensure_schema()


def require_user_id():
    value = request.headers.get('X-User-Id')
    if not value and request.is_json:
        body = request.json or {}
        value = body.get('user_id')
    if not value:
        value = request.args.get('user_id')
    try:
        uid = int(value)
    except Exception:
        uid = 0
    return uid if uid > 0 else None


def hash_password(raw_password):
    return hashlib.sha256((raw_password or '').encode('utf-8')).hexdigest()


def is_valid_email(email):
    if not email:
        return False
    return re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email) is not None


def get_env_any(*names):
    for name in names:
        value = os.getenv(name)
        if value:
            return value
    return ''


def mask_secret(value):
    if not value:
        return '(missing)'
    if len(value) <= 8:
        return '***'
    return f'{value[:4]}...{value[-4:]}'


def call_qwen(messages, model='qwen-plus'):
    api_key = get_env_any('QWEN_API_KEY')
    if not api_key:
        return None
    body = {
        'model': model,
        'messages': messages,
        'temperature': 0.2,
    }
    data = json.dumps(body).encode('utf-8')
    req = urllib.request.Request(
        'https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions',
        data=data,
        headers={'Content-Type': 'application/json', 'Authorization': f'Bearer {api_key}'},
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=25) as resp:
            payload = json.loads(resp.read().decode('utf-8'))
        choices = payload.get('choices') or []
        if not choices:
            return None
        return str((choices[0].get('message') or {}).get('content') or '').strip()
    except Exception:
        return None


def translate_with_qwen(english_text):
    messages = [
        {
            'role': 'system',
            'content': '请将下面英文翻译成自然流畅的中文，适合英语学习者阅读。只返回中文翻译，不要解释。',
        },
        {'role': 'user', 'content': english_text},
    ]
    return call_qwen(messages, model='qwen-plus')


def speaking_chat_with_qwen(user_text):
    prompt = (
        '你是一名英语口语陪练老师。用户会用英文和你对话。'
        '请先用自然、简短的英文回复用户，模拟真实口语交流。'
        '如果用户表达有明显语法错误或不自然的地方，请在英文回复后，用中文简短指出问题，并给出更自然的表达。'
        '如果用户表达基本正确，就简单鼓励并继续追问一个英文问题，引导对话继续。'
        f'\n\n用户说：\n{user_text}'
    )
    api_key = get_env_any('DASHSCOPE_API_KEY', 'QWEN_API_KEY')
    if not api_key:
        return None
    body = {
        'model': 'qwen-plus',
        'messages': [
            {'role': 'system', 'content': '你是英语口语陪练老师，回复自然友好，便于英语学习者练习。'},
            {'role': 'user', 'content': prompt},
        ],
        'temperature': 0.6,
    }
    req = urllib.request.Request(
        'https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions',
        data=json.dumps(body).encode('utf-8'),
        headers={'Content-Type': 'application/json', 'Authorization': f'Bearer {api_key}'},
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=25) as resp:
            payload = json.loads(resp.read().decode('utf-8'))
        choices = payload.get('choices') or []
        if not choices:
            return None
        return str((choices[0].get('message') or {}).get('content') or '').strip()
    except Exception:
        return None


def speaking_tts_with_qwen(text):
    api_key = get_env_any('DASHSCOPE_API_KEY', 'QWEN_API_KEY')
    if not api_key:
        return None, 'missing_api_key'
    body = {
        'model': 'qwen-tts',
        'input': {'text': text},
        'parameters': {
            'voice': 'Cherry',
            'format': 'mp3',
        },
    }
    req = urllib.request.Request(
        'https://dashscope.aliyuncs.com/api/v1/services/aigc/multimodal-generation/generation',
        data=json.dumps(body).encode('utf-8'),
        headers={'Content-Type': 'application/json', 'Authorization': f'Bearer {api_key}'},
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            payload = json.loads(resp.read().decode('utf-8'))
        output = payload.get('output') or {}
        audio = output.get('audio') or {}
        audio_url = audio.get('url') or ''
        audio_b64 = audio.get('data') or ''
        if audio_url:
            return {'audioUrl': audio_url, 'mime': 'audio/mpeg'}, None
        if audio_b64:
            return {'audioBase64': audio_b64, 'mime': 'audio/mpeg'}, None
        return None, 'no_audio_in_response'
    except Exception as e:
        return None, str(e)


def parse_qwen_bilingual(text):
    content = str(text or '').strip()
    if not content:
        return '', ''
    lines = [line.strip() for line in content.splitlines() if line.strip()]
    english = ''
    chinese = ''
    for line in lines:
        if line.lower().startswith('english:') or line.startswith('英文：'):
            english = line.split(':', 1)[-1].strip() if ':' in line else line.replace('英文：', '').strip()
        elif line.lower().startswith('chinese:') or line.startswith('中文：'):
            chinese = line.split(':', 1)[-1].strip() if ':' in line else line.replace('中文：', '').strip()
    placeholders = {'<英文原文>', '<中文翻译>', '英文原文', '中文翻译'}
    if english in placeholders:
        english = ''
    if chinese in placeholders:
        chinese = ''
    if english and chinese:
        return english, chinese
    # fallback: try markdown sections
    marker_en = 'English'
    marker_zh = 'Chinese'
    if marker_en in content and marker_zh in content:
        try:
            en_part = content.split(marker_en, 1)[1].split(marker_zh, 1)[0]
            zh_part = content.split(marker_zh, 1)[1]
            return en_part.replace(':', '').strip(), zh_part.replace(':', '').strip()
        except Exception:
            pass
    return '', ''


def parse_qwen_json_bilingual(text):
    content = str(text or '').strip()
    if not content:
        return '', ''
    try:
        obj = json.loads(content)
        english = str(obj.get('english') or '').strip()
        chinese = str(obj.get('chinese') or '').strip()
        return english, chinese
    except Exception:
        return '', ''


def extract_english_fallback(text):
    content = str(text or '')
    lines = [line.strip() for line in content.splitlines() if line.strip()]
    english_lines = []
    for line in lines:
        alpha = sum(ch.isalpha() for ch in line)
        if alpha >= 3:
            english_lines.append(line)
    return '\n'.join(english_lines).strip()


def enrich_dictionary_fields(word, definition, example, example_zh, collocations):
    need_example = not str(example or '').strip()
    need_example_zh = not str(example_zh or '').strip()
    need_collocations = not isinstance(collocations, list) or len(collocations) == 0
    if not (need_example or need_example_zh or need_collocations):
        return example, example_zh, collocations

    prompt = (
        "你是英语学习助手。请仅返回 JSON，不要解释。"
        "字段: example_en(英文例句), example_zh(中文翻译), collocations(英文常见搭配数组, 3-5个)。"
        f"单词: {word}\n"
        f"英文释义: {definition or ''}\n"
        "要求: 内容简短自然，搭配只要短语。"
    )
    raw = call_qwen(
        [
            {'role': 'system', 'content': 'You are a helpful English learning assistant.'},
            {'role': 'user', 'content': prompt},
        ],
        model='qwen-plus',
    )
    if not raw:
        return example, example_zh, collocations

    payload = None
    try:
        payload = json.loads(raw)
    except Exception:
        m = re.search(r'\{[\s\S]*\}', raw)
        if m:
            try:
                payload = json.loads(m.group(0))
            except Exception:
                payload = None
    if not isinstance(payload, dict):
        return example, example_zh, collocations

    ai_example = str(payload.get('example_en') or '').strip()
    ai_example_zh = str(payload.get('example_zh') or '').strip()
    ai_collocations = payload.get('collocations')
    if not isinstance(ai_collocations, list):
        ai_collocations = []
    ai_collocations = [str(item).strip() for item in ai_collocations if str(item).strip()]

    final_example = example or ai_example
    final_example_zh = example_zh or ai_example_zh
    final_collocations = collocations if isinstance(collocations, list) and len(collocations) > 0 else ai_collocations
    return final_example, final_example_zh, final_collocations


def extract_local_dictionary_fallback(entry):
    example = ''
    collocations = []
    if not isinstance(entry, dict):
        return example, collocations
    meanings = entry.get('meanings') or []
    if not isinstance(meanings, list):
        meanings = []
    for meaning in meanings:
        if not isinstance(meaning, dict):
            continue
        defs = (meaning or {}).get('definitions') or []
        if not isinstance(defs, list):
            defs = []
        for d in defs:
            if not isinstance(d, dict):
                continue
            ex = str((d or {}).get('example') or '').strip()
            if ex and not example:
                example = ex
            # 用同义词/反义词和释义短语做基础搭配兜底
            for key in ('synonyms', 'antonyms'):
                values = (d or {}).get(key) or []
                if isinstance(values, list):
                    for v in values[:3]:
                        text = str(v or '').strip()
                        if text:
                            collocations.append(text)
            definition = str((d or {}).get('definition') or '').strip()
            if definition:
                # 取释义中的前半句作为短语兜底，避免全空
                short = re.split(r'[.;:]', definition)[0].strip()
                if 2 <= len(short.split()) <= 6:
                    collocations.append(short)
    # 去重且限制数量
    dedup = []
    seen = set()
    for item in collocations:
        key = item.lower()
        if key in seen:
            continue
        seen.add(key)
        dedup.append(item)
        if len(dedup) >= 5:
            break
    return example, dedup


def image_to_data_url(image_bytes, filename='upload.png'):
    ext = (filename.rsplit('.', 1)[-1] if '.' in filename else 'png').lower()
    if ext == 'jpg':
        ext = 'jpeg'
    if ext not in {'png', 'jpeg', 'webp', 'bmp', 'gif'}:
        ext = 'png'
    encoded = base64.b64encode(image_bytes).decode('utf-8')
    return f'data:image/{ext};base64,{encoded}'


@app.route('/api/readings', methods=['GET'])
def list_readings():
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute(
        '''
        SELECT id, title, category, COALESCE(summary, '') AS summary, english, chinese,
               COALESCE(NULLIF(TRIM(date), ''), DATE(created_at)) AS date
        FROM readings
        WHERE user_id = ?
        ORDER BY id DESC
        ''',
        (user_id,),
    )
    rows = cursor.fetchall()
    conn.close()
    return jsonify(
        [
            {
                'id': r['id'],
                'title': r['title'],
                'category': r['category'],
                'summary': r['summary'],
                'english': r['english'],
                'chinese': r['chinese'],
                'date': r['date'],
            }
            for r in rows
        ]
    )


@app.route('/api/readings/<int:reading_id>', methods=['GET'])
def get_reading_detail(reading_id):
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute(
        '''
        SELECT id, title, category, COALESCE(summary, '') AS summary, english, chinese,
               COALESCE(NULLIF(TRIM(date), ''), DATE(created_at)) AS date
        FROM readings
        WHERE id = ? AND user_id = ?
        LIMIT 1
        ''',
        (reading_id, user_id),
    )
    row = cursor.fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'not found'}), 404
    return jsonify(
        {
            'id': row['id'],
            'title': row['title'],
            'category': row['category'],
            'summary': row['summary'],
            'english': row['english'],
            'chinese': row['chinese'],
            'date': row['date'],
        }
    )


@app.route('/api/readings', methods=['POST'])
def create_reading():
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    data = request.json or {}
    title = (data.get('title') or '').strip() or 'AI导入文章'
    category = (data.get('category') or '').strip() or '我的导入'
    english = (data.get('english') or '').strip()
    chinese = (data.get('chinese') or '').strip()
    summary = (data.get('summary') or '').strip()
    if not english or not chinese:
        return jsonify({'error': 'english and chinese are required'}), 400
    if not summary:
        summary = english[:100].replace('\n', ' ')
    today = datetime.now().strftime('%Y-%m-%d')

    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute(
        '''
        INSERT INTO readings (user_id, title, category, summary, english, chinese, date)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ''',
        (user_id, title, category, summary, english, chinese, today),
    )
    reading_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': reading_id, 'title': title, 'category': category, 'english': english, 'chinese': chinese})


@app.route('/api/readings/batch_delete', methods=['POST'])
def batch_delete_readings():
    user_id = require_user_id()
    if not user_id:
        return jsonify({'success': False, 'error': 'user_id required'}), 401
    data = request.json or {}
    ids = data.get('ids') or []
    if not isinstance(ids, list):
        return jsonify({'success': False, 'error': 'ids must be a list'}), 400

    normalized_ids = []
    for item in ids:
        try:
            value = int(item)
            if value > 0:
                normalized_ids.append(value)
        except Exception:
            continue
    normalized_ids = list(dict.fromkeys(normalized_ids))

    if len(normalized_ids) == 0:
        return jsonify({'success': False, 'error': '没有可删除文章', 'deletedCount': 0}), 400

    conn = get_conn()
    cursor = conn.cursor()
    placeholders = ','.join(['?'] * len(normalized_ids))
    cursor.execute(
        f'SELECT COUNT(*) AS c FROM readings WHERE user_id = ? AND id IN ({placeholders})',
        (user_id, *normalized_ids),
    )
    matched = cursor.fetchone()['c'] or 0
    if matched == 0:
        conn.close()
        return jsonify({'success': False, 'error': '没有可删除文章', 'deletedCount': 0}), 404

    cursor.execute(
        f'DELETE FROM readings WHERE user_id = ? AND id IN ({placeholders})',
        (user_id, *normalized_ids),
    )
    deleted_count = cursor.rowcount if cursor.rowcount is not None else 0
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'deletedCount': int(deleted_count), 'matchedCount': int(matched)})


@app.route('/api/ai/reading/import', methods=['POST'])
def ai_import_reading():
    english = ''
    chinese = ''
    if request.files.get('image') or request.files.get('file'):
        image = request.files.get('image') or request.files.get('file')
        image_bytes = image.read() if image else b''
        if not image_bytes:
            return jsonify({'success': False, 'error': '上传图片为空，请重新选择图片'}), 400
        data_url = image_to_data_url(image_bytes, getattr(image, 'filename', 'upload.png'))
        prompt = (
            '请识别图片中的英文文本。'
            '只返回英文原文，不要翻译，不要解释，不要加标签。'
            '如果图片里没有可读英文，返回空字符串。'
        )
        raw = call_qwen(
            [
                {'role': 'system', 'content': '你是英语阅读助手，负责图片英文识别与中译。'},
                {
                    'role': 'user',
                    'content': [
                        {'type': 'text', 'text': prompt},
                        {'type': 'image_url', 'image_url': {'url': data_url}},
                    ],
                },
            ],
            model='qwen-vl-plus'
        )
        if not raw:
            return jsonify({'success': False, 'error': '翻译失败，请稍后重试'}), 500
        print(f"[AI_IMPORT] qwen-vl raw: {str(raw)[:300]}")
        english = extract_english_fallback(raw)
        if not english:
            return jsonify({'success': False, 'error': '未识别到有效英文内容'}), 400
        chinese = translate_with_qwen(english) or ''
        if not chinese:
            return jsonify({'success': False, 'error': '翻译失败，请稍后重试'}), 500
    if not english:
        body = request.json if request.is_json else request.form
        english = str((body or {}).get('text') or '').strip()
    if not english:
        return jsonify({'success': False, 'error': '请上传图片或输入英文文本'}), 400
    if len(english) < 2:
        return jsonify({'success': False, 'error': '未识别到有效英文内容'}), 400

    if not chinese:
        chinese = translate_with_qwen(english)
    if not chinese:
        return jsonify({'success': False, 'error': '翻译失败，请稍后重试'}), 500

    return jsonify({'success': True, 'english': english, 'chinese': chinese})


@app.route('/auth/register', methods=['POST'])
def register_user():
    data = request.json or {}
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    email = (data.get('email') or '').strip().lower()
    avatar = (data.get('avatar') or '').strip()

    if len(username) < 2:
        return jsonify({'error': '用户名至少 2 位'}), 400
    if len(password) < 6:
        return jsonify({'error': '密码至少 6 位'}), 400
    if not is_valid_email(email):
        return jsonify({'error': '邮箱格式不正确'}), 400

    if len(avatar) > 1_200_000:
        return jsonify({'error': '头像过大'}), 400

    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM users WHERE username = ? LIMIT 1', (username,))
    if cursor.fetchone():
        conn.close()
        return jsonify({'error': '用户名已存在'}), 400
    cursor.execute('SELECT id FROM users WHERE email = ? LIMIT 1', (email,))
    if cursor.fetchone():
        conn.close()
        return jsonify({'error': '邮箱已注册'}), 400

    cursor.execute(
        '''
        INSERT INTO users (username, email, avatar, password_hash)
        VALUES (?, ?, ?, ?)
        ''',
        (username, email, avatar, hash_password(password)),
    )
    user_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify(
        {'msg': 'registered', 'user': {'id': user_id, 'username': username, 'email': email, 'avatar': avatar}}
    )


@app.route('/auth/login', methods=['POST'])
def login_user():
    data = request.json or {}
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    if not username or not password:
        return jsonify({'error': 'username and password are required'}), 400

    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute(
        '''
        SELECT id, username, email, COALESCE(avatar, '') AS avatar, password_hash
        FROM users
        WHERE username = ?
        LIMIT 1
        ''',
        (username,),
    )
    row = cursor.fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'user not found'}), 404
    if row['password_hash'] != hash_password(password):
        return jsonify({'error': 'password incorrect'}), 401
    return jsonify(
        {
            'msg': 'ok',
            'user': {'id': row['id'], 'username': row['username'], 'email': row['email'], 'avatar': row['avatar']},
        }
    )


@app.route('/books', methods=['GET'])
def get_books():
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute(
        '''
        SELECT b.id, b.name, b.description, b.cover, COUNT(w.id) AS total_count
        , COALESCE(NULLIF(TRIM(b.tag), ''), '未分类') AS tag
        FROM books b
        LEFT JOIN words w ON w.book_id = b.id AND w.user_id = b.user_id
        WHERE b.user_id = ?
        GROUP BY b.id
        ORDER BY b.id DESC
        ''',
        (user_id,),
    )
    rows = cursor.fetchall()
    conn.close()
    return jsonify(
        [
            {
                'id': r['id'],
                'name': r['name'],
                'description': r['description'],
                'cover': r['cover'],
                'tag': r['tag'],
                'totalCount': r['total_count'] or 0,
            }
            for r in rows
        ]
    )


@app.route('/books', methods=['POST'])
def create_book():
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    data = request.json or {}
    name = (data.get('name') or '').strip()
    tag = (data.get('tag') or '').strip()
    description = (data.get('description') or '').strip()
    cover = (data.get('cover') or '').strip()
    if not name:
        return jsonify({'error': 'name is required'}), 400

    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute(
        '''
        INSERT INTO books (user_id, name, tag, description, cover)
        VALUES (?, ?, ?, ?, ?)
        ''',
        (user_id, name, tag, description, cover),
    )
    new_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': new_id, 'name': name, 'tag': tag, 'description': description, 'cover': cover})


@app.route('/books/<int:book_id>', methods=['DELETE'])
def delete_book(book_id):
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM books WHERE user_id = ? ORDER BY id ASC", (user_id,))
    ids = [row['id'] for row in cursor.fetchall()]
    if book_id not in ids:
        conn.close()
        return jsonify({'error': 'book not found'}), 404
    if len(ids) <= 1:
        conn.close()
        return jsonify({'error': '至少保留一个词库'}), 400

    cursor.execute("DELETE FROM words WHERE book_id = ? AND user_id = ?", (book_id, user_id))
    cursor.execute("DELETE FROM books WHERE id = ? AND user_id = ?", (book_id, user_id))
    conn.commit()
    conn.close()
    return jsonify({'msg': 'deleted'})


@app.route('/books/<int:book_id>/entries', methods=['GET'])
def get_book_entries(book_id):
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    keyword = (request.args.get('keyword') or '').strip()
    conn = get_conn()
    cursor = conn.cursor()
    if keyword:
        cursor.execute(
            '''
            SELECT id, word, meaning, COALESCE(NULLIF(TRIM(pos), ''), '未知') AS pos, COALESCE(phonetic, '') AS phonetic, COALESCE(audio, '') AS audio
            FROM words
            WHERE user_id = ? AND book_id = ? AND (word LIKE ? OR meaning LIKE ?)
            ORDER BY id DESC
            ''',
            (user_id, book_id, f'%{keyword}%', f'%{keyword}%'),
        )
    else:
        cursor.execute(
            '''
            SELECT id, word, meaning, COALESCE(NULLIF(TRIM(pos), ''), '未知') AS pos, COALESCE(phonetic, '') AS phonetic, COALESCE(audio, '') AS audio
            FROM words
            WHERE user_id = ? AND book_id = ?
            ORDER BY id DESC
            ''',
            (user_id, book_id),
        )
    rows = cursor.fetchall()
    conn.close()
    return jsonify(
        {
            'bookId': book_id,
            'count': len(rows),
            'words': [
                {
                    'id': row['id'],
                    'word': row['word'],
                    'meaning': row['meaning'],
                    'pos': row['pos'],
                    'phonetic': row['phonetic'],
                    'audio': row['audio'],
                }
                for row in rows
            ],
        }
    )


@app.route('/books/<int:book_id>/entries', methods=['POST'])
def create_book_entry(book_id):
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    data = request.json or {}
    word = (data.get('word') or '').strip()
    meaning = (data.get('meaning') or '').strip()
    pos = (data.get('pos') or '未知').strip() or '未知'
    phonetic = (data.get('phonetic') or '').strip()
    audio = (data.get('audio') or '').strip()
    if not word or not meaning:
        return jsonify({'error': 'word and meaning are required'}), 400

    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute(
        '''
        INSERT INTO words (user_id, word, meaning, pos, phonetic, audio, book_id)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ''',
        (user_id, word, meaning, pos, phonetic, audio, book_id),
    )
    new_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': new_id, 'word': word, 'meaning': meaning, 'pos': pos, 'phonetic': phonetic})


@app.route('/books/<int:book_id>/entries/<int:entry_id>', methods=['DELETE'])
def delete_book_entry(book_id, entry_id):
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM words WHERE id = ? AND book_id = ? AND user_id = ?", (entry_id, book_id, user_id))
    conn.commit()
    conn.close()
    return jsonify({'msg': 'deleted'})


@app.route('/books/<int:book_id>/import_wrong', methods=['POST'])
def import_wrong_words_to_book(book_id):
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    data = request.json or {}
    items = data.get('items') or []
    if not isinstance(items, list) or len(items) == 0:
        return jsonify({'error': 'items is required'}), 400

    payload = []
    for item in items:
        word = (item.get('word') or '').strip() if isinstance(item, dict) else ''
        meaning = (item.get('meaning') or '').strip() if isinstance(item, dict) else ''
        pos = (item.get('pos') or '未知').strip() if isinstance(item, dict) else '未知'
        if not word or not meaning:
            continue
        payload.append((user_id, word, meaning, pos or '未知', '', '', book_id))

    if len(payload) == 0:
        return jsonify({'error': 'no valid items'}), 400

    conn = get_conn()
    cursor = conn.cursor()
    cursor.executemany(
        '''
        INSERT INTO words (user_id, word, meaning, pos, phonetic, audio, book_id)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ''',
        payload,
    )
    conn.commit()
    conn.close()
    return jsonify({'msg': 'imported', 'count': len(payload)})


@app.route('/books/<int:book_id>/upload_excel', methods=['POST'])
def upload_book_excel(book_id):
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    if load_workbook is None:
        return jsonify({'error': 'openpyxl is required for excel upload'}), 500

    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'file is required'}), 400

    try:
        workbook = load_workbook(filename=BytesIO(file.read()), read_only=True, data_only=True)
        sheet = workbook.active
    except Exception:
        return jsonify({'error': 'invalid excel file'}), 400

    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) == 0:
        return jsonify({'error': 'excel has no data rows'}), 400

    payload = []
    # Detect whether first row is header; if not, treat all rows as data.
    header_cells = [str(cell or '').strip().lower() for cell in (rows[0] or ())]
    has_header = False
    if len(header_cells) >= 2:
        has_header = (
            ('word' in header_cells[0] or '单词' in header_cells[0])
            and ('meaning' in header_cells[1] or '定义' in header_cells[1] or '释义' in header_cells[1])
        )

    data_rows = rows[1:] if has_header else rows

    for row in data_rows:
        if not row:
            continue
        word = str(row[0] or '').strip() if len(row) > 0 else ''
        meaning = str(row[1] or '').strip() if len(row) > 1 else ''
        phonetic = str(row[2] or '').strip() if len(row) > 2 else ''
        audio = str(row[3] or '').strip() if len(row) > 3 else ''
        if not word or not meaning:
            continue
        payload.append((user_id, word, meaning, '未知', phonetic, audio, book_id))

    if len(payload) == 0:
        return jsonify({'error': 'no valid words found; check columns (word, meaning, phonetic, audio)'}), 400

    conn = get_conn()
    cursor = conn.cursor()
    cursor.executemany(
        '''
        INSERT INTO words (user_id, word, meaning, pos, phonetic, audio, book_id)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ''',
        payload,
    )
    conn.commit()
    conn.close()
    return jsonify({'msg': 'uploaded', 'count': len(payload)})


@app.route('/books/<int:book_id>/words')
def get_book_words_by_id(book_id):
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    list_type = request.args.get('type', 'learned')
    conn = get_conn()
    cursor = conn.cursor()

    if list_type == 'unlearned':
        cursor.execute(
            '''
            SELECT id, word, meaning, COALESCE(NULLIF(TRIM(pos), ''), '未知') AS pos, COALESCE(phonetic, '') AS phonetic, COALESCE(audio, '') AS audio
            FROM words
            WHERE user_id = ? AND book_id = ?
              AND word NOT IN (
                SELECT DISTINCT word
                FROM study_logs
                WHERE user_id = ? AND action IN ('learn', 'review_correct') AND book_id = ?
              )
            ORDER BY word COLLATE NOCASE
            ''',
            (user_id, book_id, user_id, book_id),
        )
    else:
        cursor.execute(
            '''
            SELECT id, word, meaning, COALESCE(NULLIF(TRIM(pos), ''), '未知') AS pos, COALESCE(phonetic, '') AS phonetic, COALESCE(audio, '') AS audio
            FROM words
            WHERE user_id = ? AND book_id = ?
              AND word IN (
                SELECT DISTINCT word
                FROM study_logs
                WHERE user_id = ? AND action IN ('learn', 'review_correct') AND book_id = ?
              )
            ORDER BY word COLLATE NOCASE
            ''',
            (user_id, book_id, user_id, book_id),
        )

    rows = cursor.fetchall()
    conn.close()
    return jsonify(
        {
            'bookId': book_id,
            'type': list_type,
            'count': len(rows),
            'words': [
                {
                    'id': r['id'],
                    'word': r['word'],
                    'meaning': r['meaning'],
                    'pos': r['pos'],
                    'phonetic': r['phonetic'],
                    'audio': r['audio'],
                }
                for r in rows
            ],
        }
    )


@app.route('/question')
def get_question():
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    book_id = request.args.get('book_id', type=int)
    conn = get_conn()
    cursor = conn.cursor()

    if book_id:
        cursor.execute(
            '''
            SELECT word, meaning, COALESCE(NULLIF(TRIM(pos), ''), '未知') AS pos, COALESCE(NULLIF(TRIM(phonetic), ''), '') AS phonetic, COALESCE(NULLIF(TRIM(audio), ''), '') AS audio
            FROM words
            WHERE user_id = ? AND book_id = ?
            ORDER BY RANDOM()
            LIMIT 4
            ''',
            (user_id, book_id),
        )
    else:
        cursor.execute(
            '''
            SELECT word, meaning, COALESCE(NULLIF(TRIM(pos), ''), '未知') AS pos, COALESCE(NULLIF(TRIM(phonetic), ''), '') AS phonetic, COALESCE(NULLIF(TRIM(audio), ''), '') AS audio
            FROM words
            WHERE user_id = ?
            ORDER BY RANDOM()
            LIMIT 4
            ''',
            (user_id,),
        )

    rows = cursor.fetchall()
    if len(rows) == 0:
        conn.close()
        return jsonify({'error': 'no questions'})

    correct = random.choice(rows)
    options = [r['meaning'] for r in rows]
    conn.close()
    return jsonify(
        {
            'word': correct['word'],
            'answer': correct['meaning'],
            'meaning': correct['meaning'],
            'pos': correct['pos'],
            'phonetic': correct['phonetic'],
            'audio': correct['audio'],
            'options': options,
        }
    )


@app.route('/spelling_word')
def get_spelling_word():
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    mode = request.args.get('mode', 'normal')
    book_id = request.args.get('book_id', type=int)
    conn = get_conn()
    cursor = conn.cursor()

    if mode == 'wrong':
        cursor.execute(
            '''
            SELECT word, meaning, COALESCE(NULLIF(TRIM(pos), ''), '未知') AS pos
            FROM wrong_words
            WHERE user_id = ? AND is_mastered = 0 AND needs_review = 1
            ORDER BY RANDOM()
            LIMIT 1
            ''',
            (user_id,),
        )
    else:
        if book_id:
            cursor.execute(
                '''
                SELECT word, meaning, COALESCE(NULLIF(TRIM(pos), ''), '未知') AS pos
                FROM words
                WHERE user_id = ? AND book_id = ?
                ORDER BY RANDOM()
                LIMIT 1
                ''',
                (user_id, book_id),
            )
        else:
            cursor.execute(
                '''
                SELECT word, meaning, COALESCE(NULLIF(TRIM(pos), ''), '未知') AS pos
                FROM words
                WHERE user_id = ?
                ORDER BY RANDOM()
                LIMIT 1
                ''',
                (user_id,),
            )

    row = cursor.fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'no spelling words'})

    word = row['word']
    return jsonify(
        {
            'word': word,
            'meaning': row['meaning'],
            'pos': row['pos'],
            'pronunciationHint': f"首字母 {word[:1].upper()}，共 {len(word)} 个字母",
        }
    )


@app.route('/study/log', methods=['POST'])
def add_study_log():
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    data = request.json or {}
    word = data.get('word')
    book_id = data.get('bookId')
    action = data.get('action')
    if not word or action not in {'learn', 'review_correct'}:
        return jsonify({'error': 'word and valid action are required'}), 400

    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute(
        '''
        INSERT INTO study_logs (user_id, word, book_id, action)
        VALUES (?, ?, ?, ?)
        ''',
        (user_id, word, book_id, action),
    )
    conn.commit()
    conn.close()
    return jsonify({'msg': 'logged'})


@app.route('/study/today_summary')
def get_today_summary():
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    book_id = request.args.get('book_id', type=int)
    conn = get_conn()
    cursor = conn.cursor()

    if book_id:
        cursor.execute(
            '''
            SELECT COUNT(DISTINCT word) AS learned_count
            FROM study_logs
            WHERE user_id = ? AND action = 'learn'
              AND book_id = ?
              AND DATE(created_at, 'localtime') = DATE('now', 'localtime')
            ''',
            (user_id, book_id),
        )
        learned = cursor.fetchone()['learned_count'] or 0
        cursor.execute(
            '''
            SELECT COUNT(DISTINCT word) AS reviewed_count
            FROM study_logs
            WHERE user_id = ? AND action = 'review_correct'
              AND book_id = ?
              AND DATE(created_at, 'localtime') = DATE('now', 'localtime')
            ''',
            (user_id, book_id),
        )
        reviewed = cursor.fetchone()['reviewed_count'] or 0
        cursor.execute(
            '''
            SELECT COUNT(*) AS unreviewed_count
            FROM wrong_words ww
            WHERE ww.is_mastered = 0
              AND ww.needs_review = 1
              AND ww.user_id = ?
              AND EXISTS (
                SELECT 1
                FROM words w
                WHERE w.user_id = ? AND w.book_id = ?
                  AND w.word = ww.word
                  AND w.meaning = ww.meaning
              )
            ''',
            (user_id, user_id, book_id),
        )
        unreviewed = cursor.fetchone()['unreviewed_count'] or 0
    else:
        cursor.execute(
            '''
            SELECT COUNT(DISTINCT word) AS learned_count
            FROM study_logs
            WHERE user_id = ? AND action = 'learn' AND DATE(created_at, 'localtime') = DATE('now', 'localtime')
            ''',
            (user_id,),
        )
        learned = cursor.fetchone()['learned_count'] or 0
        cursor.execute(
            '''
            SELECT COUNT(DISTINCT word) AS reviewed_count
            FROM study_logs
            WHERE user_id = ? AND action = 'review_correct' AND DATE(created_at, 'localtime') = DATE('now', 'localtime')
            ''',
            (user_id,),
        )
        reviewed = cursor.fetchone()['reviewed_count'] or 0
        cursor.execute(
            '''
            SELECT COUNT(*) AS unreviewed_count
            FROM wrong_words
            WHERE user_id = ? AND is_mastered = 0 AND needs_review = 1
            ''',
            (user_id,),
        )
        unreviewed = cursor.fetchone()['unreviewed_count'] or 0

    conn.close()
    return jsonify({'learnedToday': learned, 'reviewedToday': reviewed, 'unreviewedWrong': unreviewed})


@app.route('/study/checkin_calendar')
def get_checkin_calendar():
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    threshold = 3

    now = datetime.now()
    target_year = year if year and year > 0 else now.year
    target_month = month if month and 1 <= month <= 12 else now.month

    y = f'{target_year:04d}'
    m = f'{target_month:02d}'

    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute(
        '''
        SELECT
          CAST(strftime('%d', created_at, 'localtime') AS INTEGER) AS day,
          COUNT(DISTINCT word) AS learned_count
        FROM study_logs
        WHERE user_id = ? AND action = 'learn'
          AND strftime('%Y', created_at, 'localtime') = ?
          AND strftime('%m', created_at, 'localtime') = ?
        GROUP BY strftime('%d', created_at, 'localtime')
        ORDER BY day ASC
        ''',
        (user_id, y, m),
    )
    rows = cursor.fetchall()
    conn.close()

    daily = {}
    checked_days = []
    for row in rows:
        day = int(row['day'] or 0)
        learned_count = int(row['learned_count'] or 0)
        if day <= 0:
            continue
        daily[str(day)] = learned_count
        if learned_count >= threshold:
            checked_days.append(day)

    return jsonify(
        {
            'year': target_year,
            'month': target_month,
            'threshold': threshold,
            'dailyLearned': daily,
            'checkedDays': checked_days,
        }
    )


@app.route('/dictionary_search')
def dictionary_search():
    word = (request.args.get('word') or '').strip()
    if not word:
        return jsonify({'error': 'word is required'}), 400

    url = f"https://api.dictionaryapi.dev/api/v2/entries/en/{urllib.parse.quote(word)}"
    payload = None
    last_error = ''
    for timeout in (8, 15):
        try:
            req = urllib.request.Request(
                url,
                headers={
                    'User-Agent': 'Mozilla/5.0',
                    'Accept': 'application/json',
                },
                method='GET',
            )
            with urllib.request.urlopen(req, timeout=timeout) as response:
                payload = json.loads(response.read().decode('utf-8'))
            break
        except urllib.error.HTTPError as http_error:
            if http_error.code == 404:
                return jsonify({'error': 'not_found'}), 404
            last_error = f'HTTPError {http_error.code}'
        except Exception as e:
            last_error = str(e)

    if payload is None:
        print(f"[dictionary_search] upstream failed for '{word}': {last_error}")
        return jsonify({'error': 'query_failed'}), 502

    if not isinstance(payload, list) or len(payload) == 0:
        return jsonify({'error': 'not_found'}), 404

    entry = payload[0] if isinstance(payload[0], dict) else {}
    phonetic = entry.get('phonetic') or ''
    uk_phonetic = ''
    us_phonetic = ''
    uk_audio = ''
    us_audio = ''
    fallback_audio = ''
    for item in (entry.get('phonetics') or []):
        item_text = (item or {}).get('text') or ''
        item_audio = (item or {}).get('audio') or ''
        if item_audio and not fallback_audio:
            fallback_audio = item_audio
        lowered = (item_audio or '').lower()
        if ('uk' in lowered or 'en-gb' in lowered) and not uk_audio:
            uk_audio = item_audio
            if item_text and not uk_phonetic:
                uk_phonetic = item_text
        if ('us' in lowered or 'en-us' in lowered) and not us_audio:
            us_audio = item_audio
            if item_text and not us_phonetic:
                us_phonetic = item_text
        if item_text and not phonetic:
            phonetic = item_text

    meanings = []
    part_of_speech_tags = []
    for meaning in (entry.get('meanings') or []):
        defs = (meaning or {}).get('definitions') or []
        pos = (meaning or {}).get('partOfSpeech') or ''
        if pos and pos not in part_of_speech_tags:
            part_of_speech_tags.append(pos)
        if len(defs) == 0:
            continue
        first_def = defs[0] or {}
        definition = first_def.get('definition') or ''
        example = first_def.get('example') or ''
        meanings.append({'partOfSpeech': pos, 'definition': definition, 'example': example})

    first_meaning = meanings[0] if meanings else {'definition': '', 'example': '', 'partOfSpeech': ''}
    collocations = []
    if first_meaning.get('example'):
        collocations.append(first_meaning['example'])

    definition = first_meaning.get('definition', '')
    example = first_meaning.get('example', '')
    example_zh = ''
    try:
        fallback_example, fallback_collocations = extract_local_dictionary_fallback(entry)
        example = example or fallback_example
        if len(collocations) == 0 and len(fallback_collocations) > 0:
            collocations = fallback_collocations
        example, example_zh, collocations = enrich_dictionary_fields(
            entry.get('word') or word,
            definition,
            example,
            example_zh,
            collocations,
        )
    except Exception as e:
        print(f"[dictionary_search] enrich failed: {e}")

    return jsonify(
        {
            'word': entry.get('word') or word,
            'phonetic': phonetic,
            'audio': fallback_audio,
            'ukPhonetic': uk_phonetic or phonetic,
            'usPhonetic': us_phonetic or phonetic,
            'ukAudio': uk_audio,
            'usAudio': us_audio,
            'meanings': meanings,
            'partOfSpeechTags': part_of_speech_tags,
            'definition': definition,
            'example': example,
            'exampleZh': example_zh,
            'collocations': collocations,
        }
    )


@app.route('/api/ai/speaking-chat', methods=['POST'])
def ai_speaking_chat():
    data = request.json or {}
    text = str(data.get('text') or '').strip()
    if not text:
        return jsonify({'success': False, 'error': '请输入或说出英文内容'}), 400
    ai_reply = speaking_chat_with_qwen(text)
    if not ai_reply:
        return jsonify({'success': False, 'error': 'AI回复失败，请稍后重试'}), 500
    return jsonify({'success': True, 'user_text': text, 'ai_reply': ai_reply})


@app.route('/api/ai/speaking-tts', methods=['POST'])
def ai_speaking_tts():
    data = request.json or {}
    text = str(data.get('text') or '').strip()
    if not text:
        return jsonify({'success': False, 'error': 'text is required'}), 400
    if len(text) > 500:
        text = text[:500]
    audio_payload, error = speaking_tts_with_qwen(text)
    if not audio_payload:
        return jsonify({'success': False, 'error': error or 'tts_failed'}), 502
    return jsonify({'success': True, **audio_payload})


@app.route('/wrong', methods=['POST'])
def add_wrong():
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    data = request.json or {}
    word = data.get('word')
    meaning = data.get('answer') or data.get('meaning')
    if not word or not meaning:
        return jsonify({'error': 'word and meaning are required'}), 400

    conn = get_conn()
    cursor = conn.cursor()
    pos = data.get('pos') or get_word_pos(cursor, word, meaning, user_id=user_id)
    cursor.execute(
        '''
        INSERT INTO wrong_words (user_id, word, meaning, pos, error_count, is_mastered, is_focus, needs_review, last_wrong_at)
        VALUES (?, ?, ?, ?, 1, 0, 0, 1, CURRENT_TIMESTAMP)
        ON CONFLICT(user_id, word, meaning) DO UPDATE SET
            pos = excluded.pos,
            error_count = wrong_words.error_count + 1,
            is_mastered = 0,
            needs_review = 1,
            last_wrong_at = CURRENT_TIMESTAMP
        ''',
        (user_id, word, meaning, pos),
    )
    conn.commit()
    conn.close()
    return jsonify({'msg': 'ok'})


@app.route('/wrong_question')
def get_wrong_question():
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute(
        '''
        SELECT
            ww.word,
            ww.meaning,
            COALESCE(NULLIF(TRIM(ww.pos), ''), '未知') AS pos,
            COALESCE((
                SELECT NULLIF(TRIM(w.phonetic), '')
                FROM words w
                WHERE w.user_id = ww.user_id AND w.word = ww.word AND w.meaning = ww.meaning
                ORDER BY w.id DESC
                LIMIT 1
            ), '') AS phonetic,
            COALESCE((
                SELECT NULLIF(TRIM(w.audio), '')
                FROM words w
                WHERE w.user_id = ww.user_id AND w.word = ww.word AND w.meaning = ww.meaning
                ORDER BY w.id DESC
                LIMIT 1
            ), '') AS audio
        FROM wrong_words
        ww
        WHERE ww.user_id = ? AND ww.is_mastered = 0 AND ww.needs_review = 1
        ORDER BY RANDOM()
        LIMIT 4
        ''',
        (user_id,),
    )
    rows = cursor.fetchall()
    if len(rows) == 0:
        conn.close()
        return jsonify({'error': 'no wrong questions'})
    correct = random.choice(rows)
    options = [r['meaning'] for r in rows]
    conn.close()
    return jsonify(
        {
            'word': correct['word'],
            'answer': correct['meaning'],
            'meaning': correct['meaning'],
            'pos': correct['pos'],
            'phonetic': correct['phonetic'],
            'audio': correct['audio'],
            'options': options,
        }
    )


@app.route('/wrong', methods=['GET'])
def get_wrong():
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute(
        '''
        SELECT
            ww.word,
            ww.meaning,
            COALESCE(NULLIF(TRIM(ww.pos), ''), '未知') AS pos,
            COALESCE((
                SELECT NULLIF(TRIM(w.audio), '')
                FROM words w
                WHERE w.user_id = ww.user_id AND w.word = ww.word AND w.meaning = ww.meaning
                ORDER BY w.id DESC
                LIMIT 1
            ), '') AS audio,
            COALESCE((
                SELECT b.name
                FROM words w
                LEFT JOIN books b ON b.id = w.book_id AND b.user_id = w.user_id
                WHERE w.user_id = ww.user_id AND w.word = ww.word AND w.meaning = ww.meaning
                ORDER BY w.id DESC
                LIMIT 1
            ), '未知词书') AS book_name,
            ww.error_count,
            ww.is_mastered,
            ww.is_focus,
            ww.needs_review,
            ww.last_wrong_at
        FROM wrong_words
        ww
        WHERE ww.user_id = ?
        ORDER BY ww.is_focus DESC, ww.error_count DESC, ww.last_wrong_at DESC
        ''',
        (user_id,),
    )
    rows = cursor.fetchall()
    conn.close()
    return jsonify([serialize_wrong(row) for row in rows])


@app.route('/wrong/summary')
def get_wrong_summary():
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute(
        '''
        SELECT COUNT(*) AS total_count,
               SUM(CASE WHEN is_mastered = 0 AND needs_review = 1 THEN 1 ELSE 0 END) AS pending_count
        FROM wrong_words
        WHERE user_id = ?
        ''',
        (user_id,),
    )
    summary = cursor.fetchone()
    cursor.execute(
        '''
        SELECT
            ww.word,
            ww.meaning,
            COALESCE(NULLIF(TRIM(ww.pos), ''), '未知') AS pos,
            COALESCE((
                SELECT NULLIF(TRIM(w.audio), '')
                FROM words w
                WHERE w.user_id = ww.user_id AND w.word = ww.word AND w.meaning = ww.meaning
                ORDER BY w.id DESC
                LIMIT 1
            ), '') AS audio,
            COALESCE((
                SELECT b.name
                FROM words w
                LEFT JOIN books b ON b.id = w.book_id AND b.user_id = w.user_id
                WHERE w.user_id = ww.user_id AND w.word = ww.word AND w.meaning = ww.meaning
                ORDER BY w.id DESC
                LIMIT 1
            ), '未知词书') AS book_name,
            ww.error_count,
            ww.is_mastered,
            ww.is_focus,
            ww.needs_review,
            ww.last_wrong_at
        FROM wrong_words ww
        WHERE ww.user_id = ?
        ORDER BY ww.error_count DESC, ww.last_wrong_at DESC
        ''',
        (user_id,),
    )
    heatmap_rows = [serialize_wrong(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify({'totalCount': summary['total_count'] or 0, 'pendingCount': summary['pending_count'] or 0, 'heatmap': heatmap_rows})


@app.route('/wrong/export_excel')
def export_wrong_excel():
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    if Workbook is None:
        return jsonify({'error': 'openpyxl is required for excel export'}), 500

    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute(
        '''
        SELECT
            ww.word,
            ww.meaning,
            ww.error_count,
            ww.is_mastered,
            ww.is_focus,
            ww.needs_review,
            COALESCE((
                SELECT b.name
                FROM words w
                LEFT JOIN books b ON b.id = w.book_id AND b.user_id = w.user_id
                WHERE w.user_id = ww.user_id AND w.word = ww.word AND w.meaning = ww.meaning
                ORDER BY w.id DESC
                LIMIT 1
            ), '未知词书') AS book_name
        FROM wrong_words ww
        WHERE ww.user_id = ?
        ORDER BY ww.is_focus DESC, ww.error_count DESC, ww.last_wrong_at DESC
        ''',
        (user_id,),
    )
    rows = cursor.fetchall()
    conn.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '错题本'
    sheet.append(['单词', '释义', '错误次数', '当前状态', '是否重点', '来源词书'])

    for row in rows:
        if row['is_mastered']:
            status = '已掌握'
        elif row['needs_review']:
            status = '待重练'
        else:
            status = '未分类'
        sheet.append(
            [
                row['word'],
                row['meaning'],
                row['error_count'],
                status,
                '是' if row['is_focus'] else '否',
                row['book_name'],
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='wrong_words.xlsx',
    )


@app.route('/wrong/remove', methods=['POST'])
def remove_wrong():
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    data = request.json or {}
    word = data.get('word')
    meaning = data.get('answer') or data.get('meaning')
    if not word or not meaning:
        return jsonify({'error': 'word and meaning are required'}), 400
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM wrong_words WHERE user_id = ? AND word = ? AND meaning = ?", (user_id, word, meaning))
    conn.commit()
    conn.close()
    return jsonify({'msg': 'deleted'})


@app.route('/wrong/toggle_focus', methods=['POST'])
def toggle_focus():
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    data = request.json or {}
    word = data.get('word')
    meaning = data.get('meaning')
    is_focus = 1 if data.get('isFocus') else 0
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE wrong_words SET is_focus = ? WHERE user_id = ? AND word = ? AND meaning = ?",
        (is_focus, user_id, word, meaning),
    )
    conn.commit()
    conn.close()
    return jsonify({'msg': 'updated'})


@app.route('/wrong/mark_mastered', methods=['POST'])
def mark_mastered():
    user_id = require_user_id()
    if not user_id:
        return jsonify({'error': 'user_id required'}), 401
    data = request.json or {}
    word = data.get('word')
    meaning = data.get('meaning')
    is_mastered = 1 if data.get('isMastered', True) else 0
    needs_review = 0 if is_mastered else 1
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE wrong_words SET is_mastered = ?, needs_review = ? WHERE user_id = ? AND word = ? AND meaning = ?",
        (is_mastered, needs_review, user_id, word, meaning),
    )
    conn.commit()
    conn.close()
    return jsonify({'msg': 'updated'})


if __name__ == '__main__':
    print(f"[ENV] QWEN_API_KEY={mask_secret(get_env_any('QWEN_API_KEY'))}")
    app.run(debug=True, port=5001)
